"""
Enhanced Excel export functionality with professional formatting.
This module creates Excel files that match the original document structure
similar to how the frontend displays different document types.
"""

import io
from datetime import datetime
from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


class ExcelStyler:
    """Handles Excel styling for professional document exports."""
    
    # Color scheme
    HEADER_BG = "1E40AF"  # Blue-700
    HEADER_TEXT = "FFFFFF"  # White
    SUBHEADER_BG = "3B82F6"  # Blue-500
    TABLE_HEADER_BG = "DBEAFE"  # Blue-100
    TABLE_HEADER_TEXT = "1E40AF"  # Blue-700
    ALT_ROW_BG = "F3F4F6"  # Gray-100
    BORDER_COLOR = "D1D5DB"  # Gray-300
    
    # Rebut specific colors
    REBUT_HEADER_BG = "4F46E5"  # Indigo-600
    REBUT_TABLE_BG = "10B981"  # Emerald-600
    
    # NPT specific colors
    NPT_HEADER_BG = "7C3AED"  # Purple-600
    NPT_TABLE_BG = "8B5CF6"  # Purple-500
    
    # Kosu specific colors
    KOSU_HEADER_BG = "F59E0B"  # Amber-500
    KOSU_TABLE_BG = "FBBF24"  # Amber-400
    
    @staticmethod
    def get_border(color="D1D5DB"):
        """Create border style."""
        side = Side(style='thin', color=color)
        return Border(left=side, right=side, top=side, bottom=side)
    
    @staticmethod
    def get_header_font():
        """Get header font style."""
        return Font(name='Arial', size=14, bold=True, color="FFFFFF")
    
    @staticmethod
    def get_subheader_font():
        """Get subheader font style."""
        return Font(name='Arial', size=12, bold=True, color="FFFFFF")
    
    @staticmethod
    def get_table_header_font():
        """Get table header font style."""
        return Font(name='Arial', size=10, bold=True, color="1E40AF")
    
    @staticmethod
    def get_data_font():
        """Get data font style."""
        return Font(name='Arial', size=10, color="1F2937")
    
    @staticmethod
    def get_center_alignment():
        """Get center alignment."""
        return Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    @staticmethod
    def get_left_alignment():
        """Get left alignment."""
        return Alignment(horizontal='left', vertical='center', wrap_text=True)


class RebutExcelExporter:
    """Export Rebut documents to styled Excel format."""
    
    def __init__(self):
        self.styler = ExcelStyler()
    
    def export_single(self, document: Dict[str, Any]) -> io.BytesIO:
        """Export a single Rebut document to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Rebut Document"
        
        # Document Title
        ws.merge_cells('A1:K1')
        title_cell = ws['A1']
        title_cell.value = "Formulaire de déclaration Rebuts"
        title_cell.font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color=self.styler.REBUT_HEADER_BG, 
                                     end_color=self.styler.REBUT_HEADER_BG, 
                                     fill_type="solid")
        title_cell.alignment = self.styler.get_center_alignment()
        ws.row_dimensions[1].height = 30
        
        # Document Metadata
        row = 2
        metadata = document.get('metadata', {})
        data = document.get('data', {})
        header = data.get('header', {})
        
        # Company info (if available)
        if header.get('company'):
            ws.merge_cells(f'A{row}:K{row}')
            company_cell = ws[f'A{row}']
            company_cell.value = header.get('company', 'TTE International - A Onetech company')
            company_cell.font = Font(name='Arial', size=12, bold=True)
            company_cell.alignment = self.styler.get_center_alignment()
            row += 1
        
        # Add blank row
        row += 1
        
        # Header Information Section
        ws.merge_cells(f'A{row}:K{row}')
        header_section_cell = ws[f'A{row}']
        header_section_cell.value = "Document Information"
        header_section_cell.font = self.styler.get_subheader_font()
        header_section_cell.fill = PatternFill(start_color=self.styler.SUBHEADER_BG, 
                                              end_color=self.styler.SUBHEADER_BG, 
                                              fill_type="solid")
        header_section_cell.alignment = self.styler.get_center_alignment()
        row += 1
        
        # Header fields in a 2-column layout
        header_info = [
            ("Filename:", metadata.get('filename', 'N/A')),
            ("Document Type:", metadata.get('document_type', 'Rebut')),
            ("Date:", header.get('date', 'N/A')),
            ("Ligne:", header.get('ligne', 'N/A')),
            ("OF Number:", header.get('of_number', 'N/A')),
            ("Mat:", header.get('mat', 'N/A')),
            ("Equipe:", header.get('equipe', 'N/A')),
            ("Visa:", header.get('visa', 'N/A')),
            ("Processed:", metadata.get('processed_at', 'N/A')),
        ]
        
        for i in range(0, len(header_info), 2):
            # First column pair
            label1, value1 = header_info[i]
            ws[f'A{row}'] = label1
            ws[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            ws[f'A{row}'].alignment = self.styler.get_left_alignment()
            ws[f'A{row}'].border = self.styler.get_border()
            
            ws.merge_cells(f'B{row}:E{row}')
            ws[f'B{row}'] = value1
            ws[f'B{row}'].font = self.styler.get_data_font()
            ws[f'B{row}'].alignment = self.styler.get_left_alignment()
            ws[f'B{row}'].border = self.styler.get_border()
            
            # Second column pair (if exists)
            if i + 1 < len(header_info):
                label2, value2 = header_info[i + 1]
                ws[f'F{row}'] = label2
                ws[f'F{row}'].font = Font(name='Arial', size=10, bold=True)
                ws[f'F{row}'].fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
                ws[f'F{row}'].alignment = self.styler.get_left_alignment()
                ws[f'F{row}'].border = self.styler.get_border()
                
                ws.merge_cells(f'G{row}:K{row}')
                ws[f'G{row}'] = value2
                ws[f'G{row}'].font = self.styler.get_data_font()
                ws[f'G{row}'].alignment = self.styler.get_left_alignment()
                ws[f'G{row}'].border = self.styler.get_border()
            
            row += 1
        
        # Add blank row
        row += 1
        
        # Items Table Section
        ws.merge_cells(f'A{row}:K{row}')
        items_header_cell = ws[f'A{row}']
        items_header_cell.value = "Items Data"
        items_header_cell.font = self.styler.get_subheader_font()
        items_header_cell.fill = PatternFill(start_color=self.styler.REBUT_TABLE_BG, 
                                            end_color=self.styler.REBUT_TABLE_BG, 
                                            fill_type="solid")
        items_header_cell.alignment = self.styler.get_center_alignment()
        row += 1
        
        # Table Headers
        items = data.get('items', [])
        if items:
            # Determine columns from first item
            columns = list(items[0].keys())
            
            # Column headers
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = col_name.replace('_', ' ').title()
                cell.font = self.styler.get_table_header_font()
                cell.fill = PatternFill(start_color=self.styler.TABLE_HEADER_BG, 
                                       end_color=self.styler.TABLE_HEADER_BG, 
                                       fill_type="solid")
                cell.alignment = self.styler.get_center_alignment()
                cell.border = self.styler.get_border()
            
            row += 1
            
            # Data rows
            for item_idx, item in enumerate(items):
                for col_idx, col_name in enumerate(columns, start=1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = item.get(col_name, '')
                    cell.font = self.styler.get_data_font()
                    cell.alignment = self.styler.get_left_alignment()
                    cell.border = self.styler.get_border()
                    
                    # Alternate row colors
                    if item_idx % 2 == 1:
                        cell.fill = PatternFill(start_color=self.styler.ALT_ROW_BG, 
                                               end_color=self.styler.ALT_ROW_BG, 
                                               fill_type="solid")
                
                row += 1
        
        # Adjust column widths
        for col in range(1, 12):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def export_multiple(self, documents: List[Dict[str, Any]]) -> io.BytesIO:
        """Export multiple Rebut documents to Excel with separate sheets."""
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        for idx, doc in enumerate(documents):
            ws = wb.create_sheet(title=f"Rebut_{idx+1}")
            
            # Similar structure as single export
            # Document Title
            ws.merge_cells('A1:K1')
            title_cell = ws['A1']
            title_cell.value = "Formulaire de déclaration Rebuts"
            title_cell.font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color=self.styler.REBUT_HEADER_BG, 
                                         end_color=self.styler.REBUT_HEADER_BG, 
                                         fill_type="solid")
            title_cell.alignment = self.styler.get_center_alignment()
            ws.row_dimensions[1].height = 30
            
            row = 3
            data = doc.get('data', {})
            header = data.get('header', {})
            items = data.get('items', [])
            
            # Quick header info
            ws[f'A{row}'] = f"Date: {header.get('date', 'N/A')}"
            ws[f'E{row}'] = f"OF: {header.get('of_number', 'N/A')}"
            row += 2
            
            # Items table
            if items:
                columns = list(items[0].keys())
                
                # Headers
                for col_idx, col_name in enumerate(columns, start=1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = col_name.replace('_', ' ').title()
                    cell.font = self.styler.get_table_header_font()
                    cell.fill = PatternFill(start_color=self.styler.TABLE_HEADER_BG, 
                                           end_color=self.styler.TABLE_HEADER_BG, 
                                           fill_type="solid")
                    cell.alignment = self.styler.get_center_alignment()
                    cell.border = self.styler.get_border()
                
                row += 1
                
                # Data
                for item in items:
                    for col_idx, col_name in enumerate(columns, start=1):
                        cell = ws.cell(row=row, column=col_idx)
                        cell.value = item.get(col_name, '')
                        cell.font = self.styler.get_data_font()
                        cell.border = self.styler.get_border()
                    row += 1
            
            # Adjust column widths
            for col in range(1, len(columns) + 1 if items else 12):
                ws.column_dimensions[get_column_letter(col)].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output


class NPTExcelExporter:
    """Export NPT documents to styled Excel format."""
    
    def __init__(self):
        self.styler = ExcelStyler()
    
    def export_single(self, document: Dict[str, Any]) -> io.BytesIO:
        """Export a single NPT document to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "NPT Document"
        
        # Document Title
        ws.merge_cells('A1:L1')
        title_cell = ws['A1']
        title_cell.value = "Non-Productive Time (NPT) Report"
        title_cell.font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color=self.styler.NPT_HEADER_BG, 
                                     end_color=self.styler.NPT_HEADER_BG, 
                                     fill_type="solid")
        title_cell.alignment = self.styler.get_center_alignment()
        ws.row_dimensions[1].height = 30
        
        # Document Metadata
        row = 3
        metadata = document.get('metadata', {})
        data = document.get('data', {})
        header = data.get('header', {})
        
        # Header Information Section
        ws.merge_cells(f'A{row}:L{row}')
        header_section_cell = ws[f'A{row}']
        header_section_cell.value = "Document Information"
        header_section_cell.font = self.styler.get_subheader_font()
        header_section_cell.fill = PatternFill(start_color=self.styler.SUBHEADER_BG, 
                                              end_color=self.styler.SUBHEADER_BG, 
                                              fill_type="solid")
        header_section_cell.alignment = self.styler.get_center_alignment()
        row += 1
        
        # Header fields
        header_info = [
            ("Filename:", metadata.get('filename', 'N/A')),
            ("Document Type:", metadata.get('document_type', 'NPT')),
            ("Date:", header.get('date', 'N/A')),
            ("UAP:", header.get('uap', 'N/A')),
            ("Equipe:", header.get('equipe', 'N/A')),
            ("Processed:", metadata.get('processed_at', 'N/A')),
        ]
        
        for i in range(0, len(header_info), 2):
            # First column pair
            label1, value1 = header_info[i]
            ws[f'A{row}'] = label1
            ws[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            ws[f'A{row}'].alignment = self.styler.get_left_alignment()
            ws[f'A{row}'].border = self.styler.get_border()
            
            ws.merge_cells(f'B{row}:E{row}')
            ws[f'B{row}'] = value1
            ws[f'B{row}'].font = self.styler.get_data_font()
            ws[f'B{row}'].alignment = self.styler.get_left_alignment()
            ws[f'B{row}'].border = self.styler.get_border()
            
            # Second column pair (if exists)
            if i + 1 < len(header_info):
                label2, value2 = header_info[i + 1]
                ws[f'F{row}'] = label2
                ws[f'F{row}'].font = Font(name='Arial', size=10, bold=True)
                ws[f'F{row}'].fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
                ws[f'F{row}'].alignment = self.styler.get_left_alignment()
                ws[f'F{row}'].border = self.styler.get_border()
                
                ws.merge_cells(f'G{row}:L{row}')
                ws[f'G{row}'] = value2
                ws[f'G{row}'].font = self.styler.get_data_font()
                ws[f'G{row}'].alignment = self.styler.get_left_alignment()
                ws[f'G{row}'].border = self.styler.get_border()
            
            row += 1
        
        # Add blank row
        row += 1
        
        # Downtime Events Table
        ws.merge_cells(f'A{row}:L{row}')
        events_header_cell = ws[f'A{row}']
        events_header_cell.value = "Downtime Events"
        events_header_cell.font = self.styler.get_subheader_font()
        events_header_cell.fill = PatternFill(start_color=self.styler.NPT_TABLE_BG, 
                                             end_color=self.styler.NPT_TABLE_BG, 
                                             fill_type="solid")
        events_header_cell.alignment = self.styler.get_center_alignment()
        row += 1
        
        # Table data
        events = data.get('downtime_events', [])
        if events:
            columns = list(events[0].keys())
            
            # Column headers
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = col_name.replace('_', ' ').title()
                cell.font = self.styler.get_table_header_font()
                cell.fill = PatternFill(start_color=self.styler.TABLE_HEADER_BG, 
                                       end_color=self.styler.TABLE_HEADER_BG, 
                                       fill_type="solid")
                cell.alignment = self.styler.get_center_alignment()
                cell.border = self.styler.get_border()
            
            row += 1
            
            # Data rows
            for event_idx, event in enumerate(events):
                for col_idx, col_name in enumerate(columns, start=1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = event.get(col_name, '')
                    cell.font = self.styler.get_data_font()
                    cell.alignment = self.styler.get_left_alignment()
                    cell.border = self.styler.get_border()
                    
                    # Alternate row colors
                    if event_idx % 2 == 1:
                        cell.fill = PatternFill(start_color=self.styler.ALT_ROW_BG, 
                                               end_color=self.styler.ALT_ROW_BG, 
                                               fill_type="solid")
                
                row += 1
        
        # Adjust column widths
        for col in range(1, 13):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def export_multiple(self, documents: List[Dict[str, Any]]) -> io.BytesIO:
        """Export multiple NPT documents to Excel with separate sheets."""
        wb = Workbook()
        wb.remove(wb.active)
        
        for idx, doc in enumerate(documents):
            ws = wb.create_sheet(title=f"NPT_{idx+1}")
            
            # Document Title
            ws.merge_cells('A1:L1')
            title_cell = ws['A1']
            title_cell.value = "Non-Productive Time (NPT) Report"
            title_cell.font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color=self.styler.NPT_HEADER_BG, 
                                         end_color=self.styler.NPT_HEADER_BG, 
                                         fill_type="solid")
            title_cell.alignment = self.styler.get_center_alignment()
            ws.row_dimensions[1].height = 30
            
            row = 3
            data = doc.get('data', {})
            header = data.get('header', {})
            events = data.get('downtime_events', [])
            
            # Quick header
            ws[f'A{row}'] = f"Date: {header.get('date', 'N/A')}"
            ws[f'E{row}'] = f"UAP: {header.get('uap', 'N/A')}"
            row += 2
            
            # Events table
            if events:
                columns = list(events[0].keys())
                
                # Headers
                for col_idx, col_name in enumerate(columns, start=1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = col_name.replace('_', ' ').title()
                    cell.font = self.styler.get_table_header_font()
                    cell.fill = PatternFill(start_color=self.styler.TABLE_HEADER_BG, 
                                           end_color=self.styler.TABLE_HEADER_BG, 
                                           fill_type="solid")
                    cell.alignment = self.styler.get_center_alignment()
                    cell.border = self.styler.get_border()
                
                row += 1
                
                # Data
                for event in events:
                    for col_idx, col_name in enumerate(columns, start=1):
                        cell = ws.cell(row=row, column=col_idx)
                        cell.value = event.get(col_name, '')
                        cell.font = self.styler.get_data_font()
                        cell.border = self.styler.get_border()
                    row += 1
            
            # Adjust column widths
            for col in range(1, len(columns) + 1 if events else 13):
                ws.column_dimensions[get_column_letter(col)].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output


class KosuExcelExporter:
    """Export Kosu documents to styled Excel format."""
    
    def __init__(self):
        self.styler = ExcelStyler()
    
    def export_single(self, document: Dict[str, Any]) -> io.BytesIO:
        """Export a single Kosu document to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Kosu Document"
        
        # Document Title
        ws.merge_cells('A1:L1')
        title_cell = ws['A1']
        title_cell.value = "Kosu Production Report"
        title_cell.font = Font(name='Arial', size=16, bold=True, color="1F2937")
        title_cell.fill = PatternFill(start_color=self.styler.KOSU_HEADER_BG, 
                                     end_color=self.styler.KOSU_HEADER_BG, 
                                     fill_type="solid")
        title_cell.alignment = self.styler.get_center_alignment()
        ws.row_dimensions[1].height = 30
        
        row = 3
        metadata = document.get('metadata', {})
        data = document.get('data', {})
        header = data.get('header', {})
        
        # Header Information
        ws.merge_cells(f'A{row}:L{row}')
        header_section_cell = ws[f'A{row}']
        header_section_cell.value = "Document Information"
        header_section_cell.font = self.styler.get_subheader_font()
        header_section_cell.fill = PatternFill(start_color=self.styler.SUBHEADER_BG, 
                                              end_color=self.styler.SUBHEADER_BG, 
                                              fill_type="solid")
        header_section_cell.alignment = self.styler.get_center_alignment()
        row += 1
        
        # Header fields
        header_info = [
            ("Filename:", metadata.get('filename', 'N/A')),
            ("Document Type:", metadata.get('document_type', 'Kosu')),
            ("Date:", header.get('date', 'N/A')),
            ("Nom Ligne:", header.get('nom_ligne', 'N/A')),
            ("Code Ligne:", header.get('code_ligne', 'N/A')),
            ("Numero OF:", header.get('numero_of', 'N/A')),
            ("Ref PF:", header.get('ref_pf', 'N/A')),
            ("Processed:", metadata.get('processed_at', 'N/A')),
        ]
        
        for i in range(0, len(header_info), 2):
            label1, value1 = header_info[i]
            ws[f'A{row}'] = label1
            ws[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            ws[f'A{row}'].alignment = self.styler.get_left_alignment()
            ws[f'A{row}'].border = self.styler.get_border()
            
            ws.merge_cells(f'B{row}:E{row}')
            ws[f'B{row}'] = value1
            ws[f'B{row}'].font = self.styler.get_data_font()
            ws[f'B{row}'].alignment = self.styler.get_left_alignment()
            ws[f'B{row}'].border = self.styler.get_border()
            
            if i + 1 < len(header_info):
                label2, value2 = header_info[i + 1]
                ws[f'F{row}'] = label2
                ws[f'F{row}'].font = Font(name='Arial', size=10, bold=True)
                ws[f'F{row}'].fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
                ws[f'F{row}'].alignment = self.styler.get_left_alignment()
                ws[f'F{row}'].border = self.styler.get_border()
                
                ws.merge_cells(f'G{row}:L{row}')
                ws[f'G{row}'] = value2
                ws[f'G{row}'].font = self.styler.get_data_font()
                ws[f'G{row}'].alignment = self.styler.get_left_alignment()
                ws[f'G{row}'].border = self.styler.get_border()
            
            row += 1
        
        row += 1
        
        # Export all data sections dynamically
        for section_key in data.keys():
            if section_key in ['header', 'document_type']:
                continue
            
            section_data = data.get(section_key)
            
            if isinstance(section_data, list) and section_data:
                # Section Header
                ws.merge_cells(f'A{row}:L{row}')
                section_header_cell = ws[f'A{row}']
                section_header_cell.value = section_key.replace('_', ' ').title()
                section_header_cell.font = self.styler.get_subheader_font()
                section_header_cell.fill = PatternFill(start_color=self.styler.KOSU_TABLE_BG, 
                                                      end_color=self.styler.KOSU_TABLE_BG, 
                                                      fill_type="solid")
                section_header_cell.alignment = self.styler.get_center_alignment()
                row += 1
                
                # Table
                columns = list(section_data[0].keys())
                
                # Headers
                for col_idx, col_name in enumerate(columns, start=1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = col_name.replace('_', ' ').title()
                    cell.font = self.styler.get_table_header_font()
                    cell.fill = PatternFill(start_color=self.styler.TABLE_HEADER_BG, 
                                           end_color=self.styler.TABLE_HEADER_BG, 
                                           fill_type="solid")
                    cell.alignment = self.styler.get_center_alignment()
                    cell.border = self.styler.get_border()
                
                row += 1
                
                # Data
                for item_idx, item in enumerate(section_data):
                    for col_idx, col_name in enumerate(columns, start=1):
                        cell = ws.cell(row=row, column=col_idx)
                        cell.value = item.get(col_name, '')
                        cell.font = self.styler.get_data_font()
                        cell.alignment = self.styler.get_left_alignment()
                        cell.border = self.styler.get_border()
                        
                        if item_idx % 2 == 1:
                            cell.fill = PatternFill(start_color=self.styler.ALT_ROW_BG, 
                                                   end_color=self.styler.ALT_ROW_BG, 
                                                   fill_type="solid")
                    
                    row += 1
                
                row += 1  # Blank row between sections
            
            elif isinstance(section_data, dict):
                # Section Header
                ws.merge_cells(f'A{row}:L{row}')
                section_header_cell = ws[f'A{row}']
                section_header_cell.value = section_key.replace('_', ' ').title()
                section_header_cell.font = self.styler.get_subheader_font()
                section_header_cell.fill = PatternFill(start_color=self.styler.KOSU_TABLE_BG, 
                                                      end_color=self.styler.KOSU_TABLE_BG, 
                                                      fill_type="solid")
                section_header_cell.alignment = self.styler.get_center_alignment()
                row += 1
                
                # Display as key-value pairs
                for key, value in section_data.items():
                    ws[f'A{row}'] = key.replace('_', ' ').title()
                    ws[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
                    ws[f'A{row}'].fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
                    ws[f'A{row}'].border = self.styler.get_border()
                    
                    ws.merge_cells(f'B{row}:L{row}')
                    ws[f'B{row}'] = value
                    ws[f'B{row}'].font = self.styler.get_data_font()
                    ws[f'B{row}'].border = self.styler.get_border()
                    
                    row += 1
                
                row += 1
        
        # Adjust column widths
        for col in range(1, 13):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def export_multiple(self, documents: List[Dict[str, Any]]) -> io.BytesIO:
        """Export multiple Kosu documents to Excel with separate sheets."""
        wb = Workbook()
        wb.remove(wb.active)
        
        for idx, doc in enumerate(documents):
            ws = wb.create_sheet(title=f"Kosu_{idx+1}")
            
            # Document Title
            ws.merge_cells('A1:L1')
            title_cell = ws['A1']
            title_cell.value = "Kosu Production Report"
            title_cell.font = Font(name='Arial', size=16, bold=True, color="1F2937")
            title_cell.fill = PatternFill(start_color=self.styler.KOSU_HEADER_BG, 
                                         end_color=self.styler.KOSU_HEADER_BG, 
                                         fill_type="solid")
            title_cell.alignment = self.styler.get_center_alignment()
            
            row = 3
            data = doc.get('data', {})
            header = data.get('header', {})
            
            # Quick header
            ws[f'A{row}'] = f"Date: {header.get('date', 'N/A')}"
            ws[f'E{row}'] = f"Ligne: {header.get('nom_ligne', 'N/A')}"
            row += 2
            
            # Export data sections
            for section_key in data.keys():
                if section_key in ['header', 'document_type']:
                    continue
                
                section_data = data.get(section_key)
                
                if isinstance(section_data, list) and section_data:
                    # Section header
                    ws.merge_cells(f'A{row}:L{row}')
                    ws[f'A{row}'] = section_key.replace('_', ' ').title()
                    ws[f'A{row}'].font = self.styler.get_subheader_font()
                    ws[f'A{row}'].fill = PatternFill(start_color=self.styler.KOSU_TABLE_BG, 
                                                     end_color=self.styler.KOSU_TABLE_BG, 
                                                     fill_type="solid")
                    ws[f'A{row}'].alignment = self.styler.get_center_alignment()
                    row += 1
                    
                    columns = list(section_data[0].keys())
                    
                    # Headers
                    for col_idx, col_name in enumerate(columns, start=1):
                        cell = ws.cell(row=row, column=col_idx)
                        cell.value = col_name.replace('_', ' ').title()
                        cell.font = self.styler.get_table_header_font()
                        cell.fill = PatternFill(start_color=self.styler.TABLE_HEADER_BG, 
                                               end_color=self.styler.TABLE_HEADER_BG, 
                                               fill_type="solid")
                        cell.alignment = self.styler.get_center_alignment()
                        cell.border = self.styler.get_border()
                    
                    row += 1
                    
                    # Data
                    for item in section_data:
                        for col_idx, col_name in enumerate(columns, start=1):
                            cell = ws.cell(row=row, column=col_idx)
                            cell.value = item.get(col_name, '')
                            cell.font = self.styler.get_data_font()
                            cell.border = self.styler.get_border()
                        row += 1
                    
                    row += 1
            
            # Adjust column widths
            for col in range(1, 13):
                ws.column_dimensions[get_column_letter(col)].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output


def export_document_to_excel(document: Dict[str, Any], doc_type: str) -> io.BytesIO:
    """
    Main function to export a single document to Excel based on type.
    
    Args:
        document: The document data dictionary
        doc_type: The document type ('Rebut', 'NPT', 'Kosu')
    
    Returns:
        BytesIO object containing the Excel file
    """
    if doc_type == 'Rebut':
        exporter = RebutExcelExporter()
        return exporter.export_single(document)
    elif doc_type == 'NPT':
        exporter = NPTExcelExporter()
        return exporter.export_single(document)
    elif doc_type == 'Kosu':
        exporter = KosuExcelExporter()
        return exporter.export_single(document)
    else:
        raise ValueError(f"Unsupported document type: {doc_type}")


def export_documents_to_excel(documents: List[Dict[str, Any]], doc_type: str) -> io.BytesIO:
    """
    Main function to export multiple documents to Excel based on type.
    
    Args:
        documents: List of document data dictionaries
        doc_type: The document type ('Rebut', 'NPT', 'Kosu')
    
    Returns:
        BytesIO object containing the Excel file with multiple sheets
    """
    if doc_type == 'Rebut':
        exporter = RebutExcelExporter()
        return exporter.export_multiple(documents)
    elif doc_type == 'NPT':
        exporter = NPTExcelExporter()
        return exporter.export_multiple(documents)
    elif doc_type == 'Kosu':
        exporter = KosuExcelExporter()
        return exporter.export_multiple(documents)
    else:
        raise ValueError(f"Unsupported document type: {doc_type}")

